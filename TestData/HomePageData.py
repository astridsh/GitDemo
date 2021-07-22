import openpyxl



class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul", "lastname":"Shetty", "gender":"Male"}, {"firstname":"Anshika", "lastname":"Shetty", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\++ ASTRID\\FACEMASK\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # TO GET ROWS
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # TO GET COLUMNS
                    # Dict["lastname"]="Shetty"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]